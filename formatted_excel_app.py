"""
Formatted Excel Builder - Streamlit Application
================================================
A tool to upload/paste tabular data and export it into a formatted Excel template.

ENHANCEMENTS ADDED:
-------------------
- Excel file upload (.xlsx, .xls) with multi-sheet selection
- Per-sheet column inclusion/exclusion controls
- Multiple output sheets in a single exported Excel file
- Sheet name sanitization and validation
- Cached file reading for performance
- Step-by-step UI flow

ASSUMPTIONS:
------------
1. Each input sheet generates one output sheet (1:1 mapping)
2. Template formatting is applied to each output sheet independently
3. Output sheet names default to input sheet names but are editable
4. Column selection happens after sheet selection
5. The template structure (rows 1-6 for metadata, row 7+ for data) applies to all output sheets

HOW TO RUN:
-----------
1. Install dependencies: pip install streamlit pandas openpyxl xlrd
2. Place your template.xlsx file in the same directory as this script
3. Run: streamlit run app.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
from datetime import date, datetime
import re

# =============================================================================
# CONFIGURATION
# =============================================================================

TEMPLATE_PATH = "Antenna for (Customer)_(Data Pull Name)_(Today's date in yyyymmdd).xlsx"
TARGET_SHEET = "Data"
DATA_START_ROW = 7
DATA_START_COL = 1

CELL_USER_INPUT = (1, 2)      # B1 - Data Pull Name (same as filename)
CELL_SERVICES = (2, 2)        # B2 - De-duplicated services
CELL_DISTRIBUTORS = (3, 2)    # B3 - De-duplicated distributors
CELL_FOOTNOTES = (4, 2)       # B4 - Selected footnotes
CELL_DATE = (5, 2)            # B5 - Today's date

# =============================================================================
# FOOTNOTES TABLE
# =============================================================================

FOOTNOTES_TABLE = {
    "Any data before January 2023": "We recommend using 24 months historical data for the most accurate trends. With further historical data, data quality is impacted by panel changes and signal loss.",
    "Any Demographic": "Data is collected at account level, not household level.",
    "Any Demographic with projected metrics": "Data is collected at account level, not household level. Antenna does not have full coverage of panelists' demographics, but relative trends are meaningful.",
    "Any metric that includes Plan Mix, Price Paid or LTV": "Antenna does not cover 100% of plans in all cases, but relative trends are meaningful.",
    "DMA data": "Antenna does not have full coverage of panelists' zip codes, and not all zip codes map to a DMA.",
    "Churn Rate": "A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "Survival Rate": "Survival Rate is defined as the percentage of new Subscribers in period 0 who remained Subscribed (and did not Cancel) in each period thereafter. Users who Cancel cannot re-enter the Survival Curve in subsequent months. Cohort Survival includes monthly and annual plans. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "M[x] Churn Rate": "M[x] Churn Rate is defined as the percentage of new Subscribers in Month 0 who Cancelled before the end of [Month x]. Users who Cancel cannot re-enter in subsequent periods. Cohort Survival includes monthly and annual plans. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors.",
    "Average Number of Subscriptions per Subscriber": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Bucketed Subscriptions per Subscriber": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Trial Conversion Rate": "The percentage of Trials who convert to paying Subscribers. A Trial is defined as 6 months or less. Trial Conversion Rate is calculated for the conversion month. Eligible users are those whose Trial is expiring in a given month.",
    "Overlap": "Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Resubscribe Rate": "12-month Resubscribe Rate is defined as the percentage of Gross Subscriber Adds who had previously Subscribed to and since Cancelled the same Service within the prior 12 months. This metric is calculated starting 3 months after a new Service and/or Distributor launch. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful.",
    "Switching Rate": "Switching Rate is defined as the percentage of users who Cancelled [Cancellation Service] and Signed-up to [Switch to Service] within 30 days. A Churn is counted when the Subscription lapses (for iTunes or when a Service sunsets) or else on the explicit cancellation date for all other Distributors. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful. This metric is lagged by 1 month.",
    "Traders": "Traders are defined as the number of users who transitioned from one Plan or Distributor in month 0 to another Plan or Distributor in month 1 while remaining subscribed to the service. Data is collected at account level, not household level. While not a perfect cross-account representation, relative trends are meaningful. This metric is lagged by 1 month.",
    "Trading Rate": "Trading is defined as the percentage of users who transitioned from one Plan or Distributor in month 0 to another Plan or Distributor in month 1 while remaining subscribed to the service.",
    "Serial Churners – Sign-ups": "Serial Churners are users who have Canceled 3 or more Premium SVOD Subscriptions in the previous 2 years. To avoid duplication, this metric is calculated on the unique user level, not the user x Service level. Data is collected at account level, not household level.",
    "Serial Churners – Subscribers": "Serial Churners are users who have Canceled 3 or more Premium SVOD Subscriptions in the previous 2 years. To avoid duplication, this metric is calculated on the unique user level, not the user x Service level. Data is collected at account level, not household level.",
    "Promotions": "Antenna does not cover 100% of promotions in all cases, but relative trends are meaningful.",
    "Price Paid": "Data reflects standard listed prices per plan and does not include promotions or other non-standard pricing. Start and End Dates are listed only when prices change within the period.",
    "Tenure": "Subscribers have a tenure of 1 in the month they Subscribe. Tenure is calculated from the point at which Antenna's reporting began. A tenure of 48 or more months is grouped as 48+ due to differences in historical data access per distributor. Tenure is calculated at the service-distributor level. When a Subscriber switches distributors but remains subscribed to the Service, Tenure will reset to 1.",
    "Win Back Rate": "Win Back Rate is defined as the percent of Cancels in a Cancel Month that then Resubscribed to the same Service within the given number of months after the Cancel.",
    "# of Lifetimes": "A Lifetime is an uninterrupted period in which a customer remained subscribed to a service since January 2021. While Lifetimes are calculated looking back to activity since January 2021, only active Subscribers since 2023 are included in the analysis.",
    "Content Leaderboard Event": "Content releases are those which drove at least 1.5x Sign-ups compared to the previous 8-week benchmark",
    "Price Increase": "Price start dates is defined as the first date Sign-ups are seen at that price in Antenna data across distributors.",
    "Daily Cancels": "Daily Cancels do not necessarily add up to Monthly Cancels due to buyers canceling on multiple days of the month."
}

# Keywords to match in column names for auto-detection
METRIC_KEYWORDS = {
    "Churn Rate": ["churn rate", "churn_rate"],
    "Survival Rate": ["survival rate", "survival_rate"],
    "M[x] Churn Rate": ["m0 churn", "m1 churn", "m2 churn", "m3 churn", "m6 churn", "m12 churn", "month churn"],
    "Trial Conversion Rate": ["trial conversion", "trial_conversion", "conversion rate"],
    "Overlap": ["overlap"],
    "Resubscribe Rate": ["resubscribe", "resub rate"],
    "Switching Rate": ["switching rate", "switch rate"],
    "Traders": ["traders"],
    "Trading Rate": ["trading rate"],
    "Serial Churners – Sign-ups": ["serial churner", "serial_churner"],
    "Serial Churners – Subscribers": ["serial churner", "serial_churner"],
    "Promotions": ["promotion"],
    "Price Paid": ["price paid", "price_paid"],
    "Tenure": ["tenure"],
    "Win Back Rate": ["win back", "winback"],
    "# of Lifetimes": ["lifetime", "lifetimes"],
    "DMA data": ["dma"],
    "Daily Cancels": ["daily cancel"],
}

# =============================================================================
# NEW: EXCEL SHEET UTILITIES
# =============================================================================

def sanitize_sheet_name(name):
    """Sanitize sheet name for Excel compatibility."""
    # Remove invalid characters
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # Truncate to 31 characters (Excel limit)
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    
    # Ensure not empty
    if not sanitized.strip():
        sanitized = "Sheet1"
    
    return sanitized.strip()


@st.cache_data(show_spinner=False)
def load_excel_sheets(_file):
    """Load all sheet names from an Excel file. Cached for performance."""
    try:
        excel_file = pd.ExcelFile(_file)
        return excel_file.sheet_names, excel_file
    except Exception as e:
        st.error(f"Error loading Excel file: {e}")
        return None, None


@st.cache_data(show_spinner=False)
def read_excel_sheet(_excel_file, sheet_name):
    """Read a specific sheet from Excel file. Cached for performance."""
    try:
        return pd.read_excel(_excel_file, sheet_name=sheet_name)
    except Exception as e:
        st.error(f"Error reading sheet '{sheet_name}': {e}")
        return None


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def try_parse_number(value, preserve_percent=False):
    """Attempt to convert a value to a number.
    
    Args:
        value: The value to parse
        preserve_percent: If True, returns (number, is_percent) tuple
    """
    if pd.isna(value):
        return (None, False) if preserve_percent else None
    if isinstance(value, (int, float)):
        return (value, False) if preserve_percent else value
    if isinstance(value, str):
        stripped = value.strip()
        is_percent = stripped.endswith('%')
        cleaned = stripped.replace(',', '').replace('$', '').replace('%', '')
        try:
            num = float(cleaned)
            # If it was already a percentage string (e.g., "9.42%"), convert to decimal
            if is_percent:
                num = num / 100
            return (num, is_percent) if preserve_percent else num
        except ValueError:
            return (None, False) if preserve_percent else None
    return (None, False) if preserve_percent else None


def try_parse_date(value):
    """Attempt to parse a value as a date."""
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%d/%m/%Y', '%d-%m-%Y',
            '%Y/%m/%d', '%b %d, %Y', '%B %d, %Y', '%d %b %Y', '%d %B %Y',
            '%m/%d/%y', '%m-%d-%y', '%Y%m%d', '%b-%y', '%b %Y', '%B %Y',
            '%m/%Y', '%m-%Y', '%Y-%m'
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def is_date_column(series):
    """Check if a column likely contains dates."""
    sample = series.dropna().head(20)
    if len(sample) == 0:
        return False
    date_count = sum(1 for v in sample if try_parse_date(v) is not None)
    return date_count / len(sample) > 0.5


def is_rate_column(col_name):
    """Check if a column represents a rate/percentage."""
    rate_keywords = ['rate', 'percent', 'pct', '%', 'ratio', 'share']
    return any(kw in str(col_name).lower() for kw in rate_keywords)


def format_column_name(col_name):
    """Format column name: remove underscores, capitalize, and hyphenate compound words."""
    formatted = str(col_name).replace('_', ' ').title()
    # Convert "Sign Ups" -> "Sign-ups", "Ad Tier" -> "Ad-Tier", etc.
    compound_words = ['Sign Ups', 'Ad Tier']
    for word in compound_words:
        if word in formatted:
            hyphenated = word.split()[0] + '-' + word.split()[1]
            formatted = formatted.replace(word, hyphenated)
    return formatted


def format_date_for_display(value):
    """Format a date value to MMM-YY for display."""
    parsed = try_parse_date(value)
    return parsed.strftime('%b-%y') if parsed else str(value) if pd.notna(value) else ""


def format_number_for_display(value, is_rate=False):
    """Format a number for display."""
    result = try_parse_number(value, preserve_percent=True)
    num, was_percent = result if result else (None, False)
    if num is None:
        return str(value) if pd.notna(value) else ""
    if is_rate or was_percent:
        # num is already in decimal form (e.g., 0.0942 for 9.42%)
        return f"{num * 100:.2f}%"
    return f"{int(num):,}" if num == int(num) else f"{num:,.0f}"


def detect_matching_footnotes(df):
    """Auto-detect which footnotes match the data columns."""
    matched = set()
    col_names_lower = [str(col).lower() for col in df.columns]
    all_text = " ".join(col_names_lower)
    
    for metric, keywords in METRIC_KEYWORDS.items():
        for keyword in keywords:
            if keyword in all_text:
                matched.add(metric)
                break
    
    # Check for demographic columns
    demo_keywords = ['age', 'gender', 'income', 'demographic', 'ethnicity', 'race']
    if any(kw in all_text for kw in demo_keywords):
        matched.add("Any Demographic")
    
    return list(matched)


# =============================================================================
# DATA PROCESSING
# =============================================================================

def load_input(uploaded_file):
    """Load data from an uploaded CSV file."""
    try:
        try:
            return pd.read_csv(uploaded_file, encoding='utf-8')
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding='latin-1')
    except Exception as e:
        st.error(f"Error loading CSV: {e}")
        return None


def parse_pasted_data(pasted_text):
    """Parse pasted text data into a DataFrame."""
    if not pasted_text or not pasted_text.strip():
        return None
    try:
        from io import StringIO
        first_line = pasted_text.strip().split('\n')[0]
        delimiter = '\t' if '\t' in first_line else ','
        return pd.read_csv(StringIO(pasted_text), delimiter=delimiter)
    except Exception as e:
        st.error(f"Error parsing data: {e}")
        return None


def validate_dataframe(df):
    """Validate DataFrame for common issues."""
    if df is None:
        return False, "No data loaded"
    if df.empty:
        return False, "DataFrame is empty"
    if df.columns.duplicated().any():
        dupes = df.columns[df.columns.duplicated()].tolist()
        return False, f"Duplicate columns: {dupes}"
    return True, None


def convert_numeric_columns(df):
    """Convert text numbers to actual numbers in the DataFrame."""
    df_converted = df.copy()
    for col in df_converted.columns:
        if is_date_column(df_converted[col]):
            continue
        converted = df_converted[col].apply(try_parse_number)
        if converted.notna().sum() > df_converted[col].notna().sum() * 0.5:
            non_null_mask = converted.notna()
            if non_null_mask.any():
                df_converted.loc[non_null_mask, col] = converted[non_null_mask]
    return df_converted


def transform_dataframe(df):
    """Transform DataFrame for display with proper formatting."""
    df_converted = convert_numeric_columns(df)
    df_display = pd.DataFrame()
    
    for col in df.columns:
        new_col = format_column_name(col)
        if is_date_column(df[col]):
            df_display[new_col] = df[col].apply(format_date_for_display)
        elif is_rate_column(col):
            df_display[new_col] = df_converted[col].apply(lambda x: format_number_for_display(x, is_rate=True))
        else:
            num_test = df_converted[col].apply(try_parse_number)
            if num_test.notna().sum() > len(df_converted[col].dropna()) * 0.5:
                df_display[new_col] = df_converted[col].apply(format_number_for_display)
            else:
                df_display[new_col] = df_converted[col].apply(lambda x: str(x) if pd.notna(x) else "")
    return df_display


def extract_unique_values(df, column_names):
    """Extract unique values from matching columns."""
    for col in df.columns:
        if col.lower() in [n.lower() for n in column_names]:
            return ", ".join(str(v) for v in df[col].dropna().unique())
    return ""


# =============================================================================
# EXCEL EXPORT - UPDATED FOR MULTI-SHEET
# =============================================================================

def write_sheet_to_workbook(ws, df, data_pull_name="", selected_footnotes=None):
    """Write a single DataFrame to a worksheet with formatting."""
    left_align = Alignment(horizontal='left')
    right_align = Alignment(horizontal='right')
    
    # B1: Data Pull Name
    ws.cell(row=CELL_USER_INPUT[0], column=CELL_USER_INPUT[1]).value = data_pull_name
    
    # B2: Services
    ws.cell(row=CELL_SERVICES[0], column=CELL_SERVICES[1]).value = extract_unique_values(df, [col for col in df.columns if "service" in col.lower()])
    
    # B3: Distributors
    ws.cell(row=CELL_DISTRIBUTORS[0], column=CELL_DISTRIBUTORS[1]).value = extract_unique_values(df, [col for col in df.columns if "distributor" in col.lower()])
    
    # B4: Selected footnotes
    if selected_footnotes:
        footnote_text = "\n\n".join([f"{FOOTNOTES_TABLE[metric]}" for metric in selected_footnotes])
        cell_b4 = ws.cell(row=CELL_FOOTNOTES[0], column=CELL_FOOTNOTES[1])
        cell_b4.value = footnote_text
        cell_b4.alignment = Alignment(wrap_text=True, vertical='top')
    
    # B5: Date
    date_cell = ws.cell(row=CELL_DATE[0], column=CELL_DATE[1])
    date_cell.value = date.today()
    date_cell.number_format = 'yyyy-mm-dd'
    
    # Convert numbers before writing
    df_converted = convert_numeric_columns(df)
    
    # Identify column types
    date_cols = [col for col in df.columns if is_date_column(df[col])]
    rate_cols = [col for col in df.columns if is_rate_column(col)]
    numeric_cols = []
    for col in df_converted.columns:
        if col not in date_cols:
            num_test = df_converted[col].apply(lambda x: try_parse_number(x, preserve_percent=True)[0] if try_parse_number(x, preserve_percent=True) else None)
            if num_test.notna().sum() > len(df_converted[col].dropna()) * 0.5:
                numeric_cols.append(col)
    
    # Track which columns should be right-aligned (numeric/rate columns)
    right_aligned_cols = set(rate_cols) | set(numeric_cols)
    
    # Write column headers with alignment matching data alignment
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=DATA_START_ROW, column=DATA_START_COL + col_idx)
        cell.value = format_column_name(col_name)
        # Right-align headers for numeric/rate columns
        if col_name in right_aligned_cols:
            cell.alignment = right_align
        else:
            cell.alignment = left_align
    
    # Write data rows
    for row_idx, row in enumerate(df_converted.itertuples(index=False), start=1):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=DATA_START_ROW + row_idx, column=DATA_START_COL + col_idx)
            col_name = df.columns[col_idx]
            
            if col_name in date_cols:
                parsed_date = try_parse_date(value)
                if parsed_date:
                    cell.value = parsed_date
                    cell.number_format = 'MMM-YY'
                else:
                    cell.value = value if pd.notna(value) else ""
                cell.alignment = left_align
                
            elif col_name in rate_cols:
                result = try_parse_number(value, preserve_percent=True)
                num, was_percent = result if result else (None, False)
                if num is not None:
                    cell.value = num  # Already in decimal form
                    cell.number_format = '0.00%'
                else:
                    cell.value = value if pd.notna(value) else ""
                cell.alignment = right_align
                
            elif col_name in numeric_cols:
                result = try_parse_number(value, preserve_percent=True)
                num, was_percent = result if result else (None, False)
                if num is not None:
                    if was_percent:
                        cell.value = num  # Already in decimal form
                        cell.number_format = '0.00%'
                    else:
                        cell.value = num
                        cell.number_format = '#,##0'
                else:
                    cell.value = value if pd.notna(value) else ""
                cell.alignment = right_align
                
            else:
                cell.value = value if pd.notna(value) else ""
                cell.alignment = left_align
    
    # Delete columns marked "Delete" (scan from right to left to preserve indices)
    cols_to_delete = []
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=DATA_START_ROW, column=col_idx).value
        if header_value and str(header_value).strip().lower() == "delete":
            cols_to_delete.append(col_idx)
    
    # Delete from right to left to avoid index shifting issues
    for col_idx in reversed(cols_to_delete):
        ws.delete_cols(col_idx)


def write_multiple_sheets_to_template(datasets_config, template_path):
    """
    Write multiple DataFrames to Excel template with multiple output sheets.
    
    Args:
        datasets_config: List of dicts with keys:
            - 'df': DataFrame to write
            - 'sheet_name': Output sheet name
            - 'data_pull_name': Data pull name for this sheet
            - 'footnotes': List of selected footnotes
        template_path: Path to the Excel template
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    try:
        wb = load_workbook(template_path)
        
        for idx, config in enumerate(datasets_config):
            df = config['df']
            sheet_name = sanitize_sheet_name(config['sheet_name'])
            data_pull_name = config.get('data_pull_name', '')
            footnotes = config.get('footnotes', [])
            
            # Create or get the worksheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # Copy template sheet structure if available, otherwise create new
                if TARGET_SHEET in wb.sheetnames and idx == 0:
                    ws = wb[TARGET_SHEET]
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
            
            # Write data to the worksheet
            write_sheet_to_workbook(ws, df, data_pull_name, footnotes)
        
        # Remove default template sheet if it exists and wasn't used
        if TARGET_SHEET in wb.sheetnames and len(datasets_config) > 0:
            if datasets_config[0]['sheet_name'] != TARGET_SHEET:
                del wb[TARGET_SHEET]
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except FileNotFoundError:
        st.error(f"Template not found: {template_path}")
        return None
    except Exception as e:
        st.error(f"Error writing template: {e}")
        return None


# =============================================================================
# UI COMPONENTS
# =============================================================================

def display_data_preview(df, df_transformed, sheet_name="", max_rows=10):
    """Display data preview."""
    title = f"📋 Data Preview: {sheet_name}" if sheet_name else "📋 Data Preview"
    st.subheader(title)
    st.dataframe(df_transformed.head(max_rows), use_container_width=True)
    if len(df) > max_rows:
        st.caption(f"Showing first {max_rows} of {len(df)} rows")


def display_data_summary(df):
    """Display data summary."""
    c1, c2, c3 = st.columns(3)
    c1.metric("Rows", len(df))
    c2.metric("Columns", len(df.columns))
    c3.metric("Cells", len(df) * len(df.columns))


# =============================================================================
# NEW: COLUMN SELECTION UI
# =============================================================================

def render_column_selector(sheet_name, df, key_prefix):
    """Render column inclusion/exclusion controls for a sheet."""
    st.markdown(f"#### 🎯 Column Selection: {sheet_name}")
    
    all_columns = df.columns.tolist()
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        selection_mode = st.radio(
            "Selection mode:",
            ["Include columns", "Exclude columns"],
            key=f"{key_prefix}_mode",
            help="Choose whether to select columns to include or exclude"
        )
    
    with col2:
        if selection_mode == "Include columns":
            selected_columns = st.multiselect(
                "Select columns to include:",
                options=all_columns,
                default=all_columns,
                key=f"{key_prefix}_include",
                help="Only selected columns will be included in the output"
            )
            
            if not selected_columns:
                st.warning("⚠️ No columns selected! Please select at least one column.")
                return None
            
            return selected_columns
        
        else:  # Exclude mode
            excluded_columns = st.multiselect(
                "Select columns to exclude:",
                options=all_columns,
                default=[],
                key=f"{key_prefix}_exclude",
                help="Selected columns will be excluded from the output"
            )
            
            included_columns = [col for col in all_columns if col not in excluded_columns]
            
            if not included_columns:
                st.warning("⚠️ All columns excluded! Please include at least one column.")
                return None
            
            return included_columns


# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    st.set_page_config(page_title="Antenna Formatted Excel Converter", page_icon="📊", layout="wide")
    st.title("📊 Antenna - Formatted Excel Converter")
    st.markdown("Transform your regular degular data into an Antenna Style Formatted Excel that even ~ Insights would approve of!")
    st.markdown("Please leave all bugs & enhancement requests in a comment on the jira ticket [here](https://antennalive.atlassian.net/browse/DA-6973)")
    
    # Initialize session state
    if 'datasets' not in st.session_state:
        st.session_state.datasets = {}
    if 'auto_footnotes' not in st.session_state:
        st.session_state.auto_footnotes = {}
    if 'selected_sheets' not in st.session_state:
        st.session_state.selected_sheets = []
    if 'excel_file' not in st.session_state:
        st.session_state.excel_file = None
    
    # =============================================================================
    # SIDEBAR: INPUT OPTIONS
    # =============================================================================
    
    with st.sidebar:
        st.header("⚙️ Step 1: Input Options")
        input_method = st.radio("Input method:", ["Upload CSV", "Upload Excel", "Paste Data"])
        
        st.divider()
        
        # CSV Upload (original functionality)
        if input_method == "Upload CSV":
            uploaded_file = st.file_uploader("Choose CSV", type=['csv'])
            if uploaded_file:
                df = load_input(uploaded_file)
                if df is not None:
                    st.session_state.datasets = {"CSV Data": df}
                    st.session_state.auto_footnotes = {"CSV Data": detect_matching_footnotes(df)}
                    st.session_state.selected_sheets = ["CSV Data"]
        
        # NEW: Excel Upload with sheet selection
        elif input_method == "Upload Excel":
            uploaded_file = st.file_uploader("Choose Excel file", type=['xlsx', 'xls'])
            if uploaded_file:
                sheet_names, excel_file = load_excel_sheets(uploaded_file)
                
                if sheet_names:
                    st.session_state.excel_file = excel_file
                    st.success(f"✅ Found {len(sheet_names)} sheet(s)")
                    
                    selected_sheets = st.multiselect(
                        "Select sheet(s) to process:",
                        options=sheet_names,
                        default=sheet_names[:1] if sheet_names else [],
                        help="Select one or more sheets to include in the output"
                    )
                    
                    if selected_sheets:
                        st.session_state.selected_sheets = selected_sheets
                        
                        # Load selected sheets
                        for sheet_name in selected_sheets:
                            if sheet_name not in st.session_state.datasets:
                                df = read_excel_sheet(excel_file, sheet_name)
                                if df is not None:
                                    st.session_state.datasets[sheet_name] = df
                                    st.session_state.auto_footnotes[sheet_name] = detect_matching_footnotes(df)
        
        # Paste Data (original functionality)
        else:
            pasted_data = st.text_area("Paste data:", height=200)
            if st.button("Parse Data", type="primary") and pasted_data:
                df = parse_pasted_data(pasted_data)
                if df is not None:
                    st.session_state.datasets = {"Pasted Data": df}
                    st.session_state.auto_footnotes = {"Pasted Data": detect_matching_footnotes(df)}
                    st.session_state.selected_sheets = ["Pasted Data"]
        
        st.divider()
        st.subheader("📝 Report Details")
        customer_name = st.text_input("Customer Name *", placeholder="e.g., Netflix")
        data_pull_name = st.text_input("Data Pull Name *", placeholder="e.g., Monthly Subscribers by Plan")
        
        if customer_name and data_pull_name:
            st.caption(f"📁 `Antenna for {customer_name}_{data_pull_name}_{date.today().strftime('%Y%m%d')}.xlsx`")
    
    # =============================================================================
    # MAIN AREA: DATA PROCESSING AND OUTPUT
    # =============================================================================
    
    if st.session_state.datasets:
        st.header("📊 Step 2: Review & Configure Data")
        
        # Track filtered datasets and their configurations
        filtered_datasets = {}
        output_configs = []
        
        # Process each selected sheet
        for idx, sheet_name in enumerate(st.session_state.selected_sheets):
            if sheet_name not in st.session_state.datasets:
                continue
                
            df_original = st.session_state.datasets[sheet_name]
            
            # Validate the dataframe
            is_valid, error_msg = validate_dataframe(df_original)
            if not is_valid:
                st.error(f"❌ Error in sheet '{sheet_name}': {error_msg}")
                continue
            
            # Create expandable section for each sheet
            with st.expander(f"📄 Sheet: **{sheet_name}**", expanded=len(st.session_state.selected_sheets) == 1):
                
                # Column selection UI
                selected_columns = render_column_selector(sheet_name, df_original, f"sheet_{idx}")
                
                if selected_columns is None:
                    continue
                
                # Filter dataframe to selected columns
                df_filtered = df_original[selected_columns].copy()
                
                # Transform for display
                df_transformed = transform_dataframe(df_filtered)
                
                # Display preview and summary
                display_data_preview(df_filtered, df_transformed, sheet_name)
                display_data_summary(df_filtered)
                
                # Show detected services and distributors
                services = extract_unique_values(df_filtered, [col for col in df_filtered.columns if "service" in col.lower()])
                distributors = extract_unique_values(df_filtered, [col for col in df_filtered.columns if "distributor" in col.lower()])
                
                if services or distributors:
                    col1, col2 = st.columns(2)
                    with col1:
                        if services:
                            st.info(f"**Detected Services:** {services}")
                    with col2:
                        if distributors:
                            st.info(f"**Detected Distributors:** {distributors}")
                
                st.divider()
                
                # Output sheet configuration
                st.markdown("#### 📝 Output Sheet Configuration")
                col_a, col_b = st.columns(2)
                
                with col_a:
                    output_sheet_name = st.text_input(
                        "Output sheet name:",
                        value=sheet_name,
                        key=f"output_name_{idx}",
                        help="Name for this sheet in the output Excel file"
                    )
                
                with col_b:
                    sheet_data_pull_name = st.text_input(
                        "Data pull name (for this sheet):",
                        value=data_pull_name if 'data_pull_name' in locals() else "",
                        key=f"data_pull_{idx}",
                        help="Will be inserted into cell B1"
                    )
                
                # Footnotes selection for this sheet
                st.markdown("#### 📝 Footnotes Selection")
                
                auto_detected = st.session_state.auto_footnotes.get(sheet_name, [])
                if auto_detected:
                    st.success(f"🔍 Auto-detected {len(auto_detected)} matching footnote(s)")
                
                selected_footnotes = st.multiselect(
                    "Select footnotes to include:",
                    options=list(FOOTNOTES_TABLE.keys()),
                    default=auto_detected,
                    key=f"footnotes_{idx}",
                    help="These footnotes will be inserted into cell B4"
                )
                
                if selected_footnotes:
                    with st.expander(f"📖 Preview {len(selected_footnotes)} selected footnote(s)"):
                        for metric in selected_footnotes:
                            st.markdown(f"**{metric}:** {FOOTNOTES_TABLE[metric]}")
                            st.markdown("---")
                
                # Store configuration
                filtered_datasets[sheet_name] = df_filtered
                output_configs.append({
                    'df': df_filtered,
                    'sheet_name': output_sheet_name,
                    'data_pull_name': sheet_data_pull_name,
                    'footnotes': selected_footnotes
                })
        
        # =============================================================================
        # GENERATE OUTPUT
        # =============================================================================
        
        if output_configs:
            st.divider()
            st.header("📥 Step 3: Generate Output")
            
            # Summary of what will be generated
            st.info(f"📊 **Ready to generate:** {len(output_configs)} sheet(s) in one Excel file")
            
            cols = st.columns(len(output_configs))
            for i, config in enumerate(output_configs):
                with cols[i]:
                    st.metric(f"Sheet {i+1}", config['sheet_name'])
            
            if not customer_name or not data_pull_name:
                st.warning("⚠️ Enter Customer Name and Data Pull Name in the sidebar to generate.")
            
            if st.button("🔄 Generate Multi-Sheet Excel", type="primary", disabled=not (customer_name and data_pull_name)):
                with st.spinner("Generating multi-sheet Excel file..."):
                    buffer = write_multiple_sheets_to_template(output_configs, TEMPLATE_PATH)
                    if buffer:
                        st.session_state.excel_buffer = buffer
                        st.session_state.output_filename = f"Antenna for {customer_name}_{data_pull_name}_{date.today().strftime('%Y%m%d')}.xlsx"
                        st.success(f"✅ Generated {len(output_configs)} sheet(s) successfully!")
            
            if st.session_state.get('excel_buffer'):
                st.download_button(
                    f"⬇️ Download: {st.session_state.output_filename}",
                    st.session_state.excel_buffer,
                    st.session_state.output_filename,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
    
    else:
        st.info("👈 Upload a CSV/Excel file or paste data to get started.")
        
        # Show footnotes reference table
        with st.expander("📚 Available Footnotes Reference"):
            footnotes_df = pd.DataFrame([
                {"Metric": k, "Footnote": v} for k, v in FOOTNOTES_TABLE.items()
            ])
            st.dataframe(footnotes_df, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()